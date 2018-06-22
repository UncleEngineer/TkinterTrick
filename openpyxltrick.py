import os
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.styles import Font, colors
#os.chdir('c:\\users\\al\\desktop')
wb = load_workbook(filename = 'example2.xlsx')
#sheet_ranges = wb['D1-D4']
#x = wb.get_sheet_names()
#print(x)
sheet = wb['Sheet1']

sheet['B2'].font = Font(name='TH Sarabun New',color=colors.RED,size=14, bold=True, italic=True)


print(sheet['E2'].value)

'''
rangecell = range(12)
for i in rangecell[1:]:
	y = sheet.cell(row=2, column=i).value
	print("I: ",i,y)
'''
countrow = sheet.max_row
print("Total row: ",countrow)


countcol = sheet.max_column
print("Total column: ",countcol)
#print(y)
#sheet['E2'].value = 'Motor'
#wb.save('example2.xlsx')


AA = get_column_letter(27)
print(AA)


test = column_index_from_string('AA')
print(test)


wb.create_sheet(title='Test My Sheet', index=0)
wb.save('test3.xlsx')

